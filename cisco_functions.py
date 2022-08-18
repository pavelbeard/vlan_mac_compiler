import json
import os
import re
import os
import socket
import textfsm
import pandas as pd
import preproc
import time
from netmiko import (
    ConnectHandler,
    NetmikoTimeoutException,
    NetmikoAuthenticationException,
    redispatch
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Fill, PatternFill
from openpyxl.utils import get_column_letter
from ntc_templates.parse import parse_output


class CiscoDataHandler:
    def get_info(self, device):
        try:
            output_path = os.path.join(os.getcwd(), "output")
            if not os.path.isdir(output_path):
                os.mkdir(output_path)

            if preproc.DEBUG():
                # для продакта
                with ConnectHandler(**device) as dev:
                    dev.enable()

                    if os.name == 'nt':
                        os.system('cls')
                    else:
                        os.system('clear')
                    
                    print('Начать работу скрипта на XL 192.168.254.28?\nY-да|N-зайти на другую железку')
                    start_script = input()

                    output = ''
                    match start_script:
                        case 'Y' | 'y':
                            print('Продолжаем...')
                        case 'N' | 'n':
                            print('1 - войти по ssh, 2 - telnet, 3 - самостоятельно пишем команду для подключения')
                            output = dev.find_prompt()
                            print(output)
                            ch = int(input())
                            match ch:
                                case 1:
                                    # ssh method
                                    command = 'ssh '

                                    # vrf
                                    print('Укажи VRF. Если нет - нажми Enter')
                                    vrf = input()
                                    
                                    if vrf == '':
                                        command += ''
                                    else:
                                        command += f'-vrf {vrf} '

                                    # username
                                    command += f"-l {dev.username} "

                                    # ip
                                    print('Введи ip:')
                                    ip = input()
                                    command += f'{ip}'

                                    # login
                                    dev.write_channel(command + '\n')
                                    time.sleep(3)

                                    output = dev.read_channel()
                                    print(output)

                                    if 'refused' in output:
                                        return False

                                case 2:
                                    # telnet method
                                    command = 'telnet '

                                    print('Введи ip:')
                                    ip = input()
                                    command += f'{ip} '

                                    # vrf
                                    print('Укажи VRF. Если нет - нажми Enter')
                                    vrf = input()
                                    if vrf == '':
                                        command += ''
                                    else:
                                        command += f'/vrf {vrf} '

                                    # login
                                    dev.write_channel(command + '\n')
                                    time.sleep(3)

                                    output = dev.read_channel()
                                    print(output)

                                    if 'refused' in output:
                                        return False

                                    if 'username' in output:
                                        dev.write_channel(dev.username + '\n')
                                        time.sleep(1)

                                        output = dev.read_channel()
                                        print(output)

                                case 3:
                                    # login ssh/telnet
                                    dev.write_channel(input() + '\n')
                                    time.sleep(3)

                                    output = dev.read_channel()
                                    print(output)

                                    if 'refused' in output:
                                        return False

                                    # if login telnet
                                    if 'username' in output:
                                        dev.write_channel(dev.username + '\n')
                                        time.sleep(1)

                                        output = dev.read_channel()
                                        print(output)

                            # pass
                            dev.write_channel(dev.password + '\n')
                            time.sleep(3)
                            output = dev.read_channel()
                            print(output)

                            # connect
                            dev.write_channel('\n')
                            time.sleep(3)
                            output = dev.read_channel()
                            print(output)

                            # done
                            redispatch(dev, device_type='cisco_ios')
                            output = dev.find_prompt()
                            print(output)

                        case _:
                            return False

                    # забираем данные из интересующей нас железки

                    show_vlan_brief = dev.send_command('show vlan brief', use_textfsm=True)
                    show_interfaces = dev.send_command('show interfaces', use_textfsm=True)
                    hostname = dev.find_prompt()
                    print(hostname)
                    
                    # преобразовываем их

                    return self.__handle_data(show_vlan_brief, show_interfaces, dev, hostname, output_path)
            else:
                # для дебага
                show_interfaces = pd.read_excel(os.getcwd() + '/output_sh_ifaces.xlsx', sheet_name='output')
                show_vlan_brief = pd.read_excel(os.getcwd() + '/output_sh_vlan_br.xlsx', sheet_name='output')
                self.__handle_data(show_vlan_brief, show_interfaces)

        except ConnectionRefusedError as cre:
            print(cre)
            return False
        except NetmikoTimeoutException as nte:
            print(nte)
            return False
        except socket.error as se:
            print(se)
            return False
        except NetmikoAuthenticationException as nae:
            print(nae)
            return False

    def __handle_data(self, *args):
        # создаем excel
        wb = Workbook()
        ws = wb.active

        # vlans and descriptions
        vlans = pd.DataFrame(args[0])
    
        df_rank_1 = vlans[['vlan_id', 'name']]

        # ports and descriptiosn
        replacements = json.loads(open(os.getcwd() + '/replacements.re', 'r').read())
        # выбираем порты по признакам: up, не loopback, не vlan
        df = pd.DataFrame(args[1])
        ports = df[
            (df.link_status == 'up') &
            (~df['interface'].str.contains('vlan.*', flags=re.IGNORECASE, regex=True)) &
            (~df['interface'].str.contains('lo.*', flags=re.IGNORECASE, regex=True))
            ]
        # меняем названия портов
        ports['interface'] = ports['interface'].replace(replacements, regex=True)
        df_rank_2 = ports[['interface', 'description']]

        # стили ячеек 1 ранга
        ws.cell(row=2, column=1, value='vlan')
        col = 1
        width = 100
        for column in df_rank_1.columns.tolist():
            ws.column_dimensions[get_column_letter(col)].width = width
            ws.cell(row=3, column=col, value=column)
            col += 1
            width -= 50

        # стили ячеек 2 ранга
        ws.row_dimensions[1].height = 180

        # заполняем 1 ранг (строки)
        rows = {}
        row = 3
        vlans_count = df_rank_1['vlan_id'].count()
        while row < vlans_count + 3:
            vlan_id = df_rank_1['vlan_id'][row - 3]
            ws.cell(column=1, row=row, value=vlan_id)
            ws.cell(column=2, row=row, value=df_rank_1['name'][row - 3])
            rows.update({vlan_id: row})
            row += 1

        # заполняем 2 ранг (столбцы)
        cols = {}
        col = 3
        ports_count = df_rank_2['interface'].count()
        df_rank_2 = df_rank_2.reset_index(drop=True)
        while col - 3 < ports_count:
            ws.cell(row=1, column=col, value=df_rank_2['description'][col - 3]) \
                .alignment = Alignment(textRotation=90)
            port = df_rank_2['interface'][col - 3]
            ws.cell(row=2, column=col, value=port)
            cols.update({port: col})
            col += 1

        # количество вланов
        ws.cell(row=2, column=2, value=f'Всего vlan: {vlans_count}/Всего активных портов: {ports_count}')

        # первое заполнение
        red_vlans = 0
        #
        
        for vlan_id in df_rank_1['vlan_id']:
            if preproc.DEBUG():
                data = args[2].send_command(f'show vlan id {vlan_id}', use_textfsm=True)
            else:
                vlan_id = 5
                data = pd.read_excel(os.getcwd() + '/output_sh_vl_id5.xlsx', sheet_name='output')

            df_list = []
            if preproc.DEBUG():
                df_list = pd.DataFrame(data)['interfaces'].tolist()[0]
            else:
                df_list = [i for i in re.split(r'\[\'|\'|\'\]', pd.DataFrame(data)['interfaces'][0]) if len(i) > 2]
                
            try:
                plist = df_list

                if len(plist) == 0:
                    ws.cell(row=int(rows.get(vlan_id)), column=2).fill = PatternFill(
                        start_color='FF0000',
                        end_color='FF0000',
                        fill_type='solid'
                    )
                    red_vlans += 1
                    raise ValueError(f'vlan {vlan_id} не содержит портов')
                
                for port in plist:
                    try:
                        r = int(rows.get(vlan_id))
                        c = cols.get(port)
                        ws.cell(row=r, column=c, value=port)
                        print(f'порт {port} во vlan {vlan_id} - есть')
                    except TypeError:
                        print(f'порт {port} во vlan {vlan_id} - в состоянии down')
                if not preproc.DEBUG():
                    break
            except ValueError as ve:
                print(ve)
                

        # второе заполнение
        cyan_vlans = 0
        orange_2_vlans = 0
        orange_vlans = 0
        xconnect = 0
        # аггрегируем порты во влане, если они там есть
        for vlan_id in df_rank_1['vlan_id']:
            try:
                data = None
                column = None
                if preproc.DEBUG():
                    try:
                        data = args[2].send_command(f'show mac address vlan {vlan_id}', use_textfsm=True)
                        column = 'destination_port'
                    except textfsm.parser.TextFSMError as fsme:
                        with open(os.getcwd() + '/NTC_TEMPLATES/cisco_ios_show_mac_address_vlan_new.textfsm') as template:
                            fsm = textfsm.TextFSM(template)
                            output = args[2].send_command(f'show mac address vlan {vlan_id}')
                            data = fsm.ParseText(output)
                            column = 5
                   
                else:
                    vlan_id = 5
                    data = pd.read_excel(os.getcwd() + '/output_sh_mac_add_vl_id5.xlsx', sheet_name='output')
                    
                pdf = []
                if preproc.DEBUG():
                    if len(data) == 0:
                        raise ValueError
                    pdf = pd.DataFrame(data)[column].tolist()
                    pdf = set([item[0] for item in pdf])
                    if len(pdf) == 0:
                        raise ValueError
                    if 'Router' in  pdf:
                        pdf.remove('Router')
                    if 'CPU' in  pdf:
                        pdf.remove('CPU')
                    print(pdf)
                else:
                    pdf = pd.DataFrame(data)[column].replace(
                        {r'\[\'|\'\]': ''}, regex=True
                    ).unique().tolist()
                    pdf = pd.DataFrame(data)[column].unique()
                    pdf.remove('Router')
                

                for port in pdf:
                    r = int(rows.get(vlan_id))
                    c = cols.get(port)
                    
                    try:
                        ws.cell(row=r, column=c).fill = PatternFill(
                            start_color='42F2F5',
                            end_color='42F2F5',
                            fill_type='solid'
                        )
                    except TypeError:
                        ws.cell(row=r, column=1).fill = PatternFill(
                            start_color='44633f',
                            end_color='44633f',
                            fill_type='solid'
                        )
                        xconnect += 1
                    except BaseException as be:
                        print(be)
                        

                if not preproc.DEBUG():
                    break

                if len(pdf) == 0:
                    print(f'vlan {vlan_id} содержит static mac-адреса')
                    ws.cell(row=int(rows.get(vlan_id)), column=1).fill = PatternFill(
                        start_color='FFC800',
                        end_color='FFC800',
                        fill_type='solid'
                    )
                    orange_vlans += 1
                else:
                    print(f'vlan {vlan_id} содержит dynamic mac-адреса')
                    cyan_vlans += 1
                    
            except ValueError as ve:
                print(ve)
                print(f'vlan {vlan_id} не содержит mac-адресов')
                ws.cell(row=int(rows.get(vlan_id)), column=1).fill = PatternFill(
                    start_color='FF7700',
                    end_color='FF7700',
                    fill_type='solid'
                )
                orange_2_vlans += 1
            except textfsm.parser.TextFSMError as fsme:
                print(fsme)
                print(f'vlan {vlan_id} не содержит mac-адресов')
                ws.cell(row=int(rows.get(vlan_id)), column=1).fill = PatternFill(
                    start_color='FF7700',
                    end_color='FF7700',
                    fill_type='solid'
                )
                orange_2_vlans += 1


        # summary
        ws.cell(row=vlans_count+5, column=1, value=f'всего vlan с dynamic и/или static mac-адресами: {cyan_vlans}').fill = PatternFill(
                        start_color='42F2F5',
                        end_color='42F2F5',
                        fill_type='solid'
                    )
        ws.cell(row=vlans_count+6, column=1, value=f'всего vlan со static mac-адресами (mac-адреса на самом роутере): {orange_vlans}').fill = PatternFill(
            start_color='FFC800',
            end_color='FFC800',
            fill_type='solid'
        )
        ws.cell(row=vlans_count+7, column=1, value=f'всего vlan без arp нагрузки: {orange_2_vlans} из них \|/').fill = PatternFill(
            start_color='FF7700',
            end_color='FF7700',
            fill_type='solid'
        )
        ws.cell(row=vlans_count+8, column=1, value=f'vlan, которые не прописаны ни на один порт: {red_vlans}').fill = PatternFill(
            start_color='FF0000',
            end_color='FF0000',
            fill_type='solid'
        )
        ws.cell(row=vlans_count+9, column=1, value=f'vlan с другими типами портов: {xconnect}').fill = PatternFill(
            start_color='44643f',
            end_color='44643f',
            fill_type='solid'
        )

        path = os.path.join(args[4], f'{args[3]}_output.xlsx')
        
        wb.save(path)
        print(f'Работа завершена. Таблица находится в {path}')
        return True

        
